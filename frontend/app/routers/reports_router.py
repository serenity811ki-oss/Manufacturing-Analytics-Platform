"""
Export & reporting endpoints.

Supports CSV (stdlib csv), Excel (openpyxl), and PDF (reportlab) exports
for production, downtime, maintenance, and quality datasets, plus a
combined "printable" OEE summary report.
"""
import csv
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.database import get_db
from app import models, auth, analytics

router = APIRouter(prefix="/api/reports", tags=["Reports & Export"])

DATASETS = {
    "production": (models.ProductionRecord, models.ProductionRecord.production_date),
    "downtime": (models.DowntimeEvent, models.DowntimeEvent.start_time),
    "maintenance": (models.MaintenanceRecord, models.MaintenanceRecord.scheduled_date),
    "quality": (models.QualityRecord, models.QualityRecord.inspection_date),
}


def _rows_for(db: Session, dataset: str, start: Optional[datetime], end: Optional[datetime]):
    model, date_col = DATASETS[dataset]
    q = db.query(model)
    if start:
        q = q.filter(date_col >= start)
    if end:
        q = q.filter(date_col <= end)
    rows = q.order_by(date_col.desc()).limit(5000).all()
    cols = [c.name for c in model.__table__.columns]
    out = []
    for r in rows:
        d = {}
        for c in cols:
            val = getattr(r, c)
            d[c] = val.value if hasattr(val, "value") else val
        out.append(d)
    return cols, out


@router.get("/export/csv/{dataset}")
def export_csv(
    dataset: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_any_role),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
):
    if dataset not in DATASETS:
        return {"error": f"Unknown dataset '{dataset}'. Choose from {list(DATASETS)}"}
    cols, rows = _rows_for(db, dataset, start_date, end_date)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols)
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={dataset}_export.csv"},
    )


@router.get("/export/excel/{dataset}")
def export_excel(
    dataset: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_any_role),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
):
    if dataset not in DATASETS:
        return {"error": f"Unknown dataset '{dataset}'. Choose from {list(DATASETS)}"}
    cols, rows = _rows_for(db, dataset, start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = dataset.capitalize()

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(cols)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([str(row[c]) if row[c] is not None else "" for c in cols])

    for col_cells in ws.columns:
        length = max(len(str(c.value)) for c in col_cells) if col_cells else 10
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={dataset}_export.xlsx"},
    )


@router.get("/export/pdf/oee-summary")
def export_pdf_oee_summary(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_any_role),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
):
    """Printable OEE summary report — one table, all active machines."""
    oee_rows = analytics.compute_oee_all_machines(db, start_date, end_date)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Manufacturing Analytics Platform", styles["Title"]))
    elements.append(Paragraph("OEE Summary Report", styles["Heading2"]))
    period = f"{start_date.date() if start_date else 'All-time'} to {end_date.date() if end_date else 'present'}"
    elements.append(Paragraph(f"Period: {period}  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.25 * inch))

    table_data = [["Machine", "Availability %", "Performance %", "Quality %", "OEE %", "Units Produced", "Downtime (min)"]]
    for r in oee_rows:
        table_data.append([
            r["machine_name"], r["availability"], r["performance"], r["quality"], r["oee"],
            r["units_produced"], r["downtime_minutes"],
        ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B0BEC5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECEFF1")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=oee_summary_report.pdf"},
    )
